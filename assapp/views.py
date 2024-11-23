from django.shortcuts import render, redirect
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from .models import Request, TotalParts, MatchedParts, UnmatchedParts
from django.conf import settings
import os
import pandas as pd
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.models import User
import io

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('request_list')  # Redirect to a home page or dashboard
        else:
            messages.error(request, 'Invalid username or password.')

    return render(request, 'assapp/login.html')

def signup_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        if username and password:
            if not User.objects.filter(username=username).exists():
                User.objects.create_user(username=username, password=password)
                messages.success(request, 'Sign up successful! Please log in.')
                return redirect('login')
            else:
                messages.error(request, 'Username already exists. Please choose another one.')
        else:
            messages.error(request, 'Invalid input. Please fill in all fields.')

    return render(request, 'assapp/signup.html')

# Forgot Password View
def forgot_password_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        if username and password:
            try:
                user = User.objects.get(username=username)
                user.set_password(password)
                user.save()
                messages.success(request, 'Password updated successfully! Please log in.')
                return redirect('login')
            except User.DoesNotExist:
                messages.error(request, 'Username not found. Please sign up first.')
        else:
            messages.error(request, 'Invalid input. Please fill in all fields.')

    return render(request, 'assapp/forgot_password.html')


def request_list2(request):
    # Fetch all requests from the database
    requests_list = Request.objects.all()

    # Set up pagination: 5 requests per page
    paginator = Paginator(requests_list, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Pass the page object to the template
    return render(request, 'assapp/request_list.html', {'page_obj': page_obj})


def request_list(request):
    # Get the search query from the request
    search_query = request.GET.get('q', '')

    # Fetch all requests from the database, filtering by search query if provided
    if search_query:
        requests_list = Request.objects.filter(
            Q(customer_name__icontains=search_query) |
            Q(customer_email__icontains=search_query) |
            Q(customer_code__icontains=search_query) |
            Q(customer_number__icontains=search_query) |
            Q(sales_contact__icontains=search_query) |
            Q(customer_comment__icontains=search_query) |
            Q(source_email__icontains=search_query) |
            Q(reference_code__icontains=search_query) |
            Q(request_type__icontains=search_query) |
            Q(request_id__icontains=search_query) |
            Q(total_parts_count__icontains=search_query) |
            Q(matched_parts_count__icontains=search_query) |
            Q(unmatched_parts_count__icontains=search_query)
        )
    else:
        requests_list = Request.objects.all()

    # Set up pagination: 5 requests per page
    paginator = Paginator(requests_list, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Pass the page object and search query to the template
    return render(request, 'assapp/request_list.html', {'page_obj': page_obj, 'search_query': search_query})



def delete_request(request, id):
    # Fetch the request instance
    request_instance = get_object_or_404(Request, id=id)
    
    # Get the `requestid_requesttype` value to filter related models
    requestid_requesttype = f"{request_instance.request_id}_{request_instance.request_type}"
    
    # Delete rows from related models
    TotalParts.objects.filter(requestid_requesttype=requestid_requesttype).delete()
    MatchedParts.objects.filter(requestid_requesttype=requestid_requesttype).delete()
    UnmatchedParts.objects.filter(requestid_requesttype=requestid_requesttype).delete()
    
    # Store the request_id for folder check before deleting the instance
    request_id = request_instance.request_id
    
    # Delete the instance in the Request model
    request_instance.delete()
    
    # Check if there are any remaining rows in the Request model with the same `request_id`
    remaining_requests = Request.objects.filter(request_id=request_id).exists()
    
    # If no rows remain, delete the media folder for this `request_id`
    if not remaining_requests:
        folder_path = os.path.join(settings.MEDIA_ROOT, request_id)
        if os.path.exists(folder_path):
            # Delete the folder and all its contents
            for root, dirs, files in os.walk(folder_path, topdown=False):
                for file in files:
                    os.remove(os.path.join(root, file))
                for dir in dirs:
                    os.rmdir(os.path.join(root, dir))
            os.rmdir(folder_path)
    
    # Redirect to the request list page or render a success message
    return redirect('request_list')

def edit_request(request, request_id):
    request_obj = get_object_or_404(Request, id=request_id)

    if request.method == 'POST':
        # Handle updating the form fields
        request_obj.customer_name = request.POST.get('customer_name')
        request_obj.customer_email = request.POST.get('customer_email')
        request_obj.customer_code = request.POST.get('customer_code')
        request_obj.request_id = request.POST.get('request_id')
        request_obj.request_type = request.POST.get('request_type')
        request_obj.reference_code = request.POST.get('reference_code')

        # Handle file upload (replace or create folder if needed)
        new_file = request.FILES.get('new_file')
        if new_file:
            # Define the directory path where files should be stored
            request_folder_path = os.path.join(settings.MEDIA_ROOT, str(request_obj.request_id))
            if not os.path.exists(request_folder_path):
                os.makedirs(request_folder_path)  # Create the folder if it doesn't exist

            # Define the full file path to save the uploaded file
            file_path = os.path.join(request_folder_path, new_file.name)

            # Replace the existing file if it exists, or save the new file
            with open(file_path, 'wb') as f:
                for chunk in new_file.chunks():
                    f.write(chunk)

            # If there's an existing source email file, update it
            request_obj.source_email = os.path.join(str(request_obj.request_id), new_file.name)
            request_obj.save()

        # Save other form updates
        request_obj.save()

        return redirect('view_request', request_id=request_id)

    return render(request, 'assapp/edit_request.html', {'request_obj': request_obj})

def edit_request1(request, request_id):
    # Retrieve the specific request object
    request_obj = get_object_or_404(Request, id=request_id)
    
    # Load existing files (if any)
    existing_files = [request_obj.source_email] if request_obj.source_email else []

    if request.method == 'POST':
        # Update fields from the submitted form data
        request_obj.customer_name = request.POST.get('customer_name', request_obj.customer_name)
        request_obj.customer_email = request.POST.get('customer_email', request_obj.customer_email)
        request_obj.customer_code = request.POST.get('customer_code', request_obj.customer_code)
        request_obj.request_id = request.POST.get('request_id', request_obj.request_id)
        request_obj.request_type = request.POST.get('request_type', request_obj.request_type)
        request_obj.reference_code = request.POST.get('reference_code', request_obj.reference_code)  # New field

        # Handle new file upload
        if 'new_file' in request.FILES:
            new_file = request.FILES['new_file']
            folder_path = os.path.join(settings.MEDIA_ROOT, str(request_obj.request_id))  # Ensure it's a string
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            file_path = os.path.join(folder_path, new_file.name)
            with open(file_path, 'wb+') as destination:
                for chunk in new_file.chunks():
                    destination.write(chunk)
            request_obj.source_email = os.path.join(str(request_obj.request_id), new_file.name)  # Save relative path

        # Save the updated request object
        request_obj.save()

        # Redirect to the request's detail view
        return redirect('view_request', request_id=request_id)

    # Render the edit form with existing data
    return render(request, 'assapp/edit_request.html', {
        'request_obj': request_obj,
        'existing_files': existing_files
    })

def delete_file(request, request_id):
    if request.method == 'POST':
        file_name = request.POST.get('file_name')
        request_obj = get_object_or_404(Request, id=request_id)
        file_path = os.path.join(settings.MEDIA_ROOT, request_obj.request_id, file_name)

        if os.path.exists(file_path):
            os.remove(file_path)  # Delete the file from the filesystem
            if request_obj.source_email == file_name:
                request_obj.source_email = None
                request_obj.save()

    return redirect('edit_request', request_id=request_id)

def view_request(request, request_id):
    # Fetch the Request instance by id
    request_instance = get_object_or_404(Request, id=request_id)
    
    # Extract the combined `request_id_request_type` from the request instance
    request_id_request_type = f"{request_instance.request_id}_{request_instance.request_type}"
    
    # Fetch related parts data using the `requestid_requesttype`
    total_parts = TotalParts.objects.filter(requestid_requesttype=request_id_request_type)
    matched_parts = MatchedParts.objects.filter(requestid_requesttype=request_id_request_type)
    unmatched_parts = UnmatchedParts.objects.filter(requestid_requesttype=request_id_request_type)

    # Pass all data to the template
    context = {
        'request': request_instance,
        'total_parts': total_parts,
        'matched_parts': matched_parts,
        'unmatched_parts': unmatched_parts
    }
    return render(request, 'assapp/view_request.html', context)

def export_to_excel(request, request_id):
    # Fetch the Request instance by id
    request_instance = get_object_or_404(Request, id=request_id)
    
    # Extract the combined `request_id_request_type` from the request instance
    request_id_request_type = f"{request_instance.request_id}_{request_instance.request_type}"
    
    # Fetch related parts data using the `requestid_requesttype`
    total_parts = TotalParts.objects.filter(requestid_requesttype=request_id_request_type)
    matched_parts = MatchedParts.objects.filter(requestid_requesttype=request_id_request_type)
    unmatched_parts = UnmatchedParts.objects.filter(requestid_requesttype=request_id_request_type)
    
    # Create an in-memory workbook
    output = io.BytesIO()
    workbook = openpyxl.Workbook()

    # Helper function to write data to sheets
    def write_to_sheet(sheet, headers, data, fields):
        # Write headers
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)
        
        # Write data rows
        for row_num, obj in enumerate(data, start=2):
            for col_num, field in enumerate(fields, start=1):
                sheet.cell(row=row_num, column=col_num, value=getattr(obj, field, ""))

    # Add Total Parts sheet
    total_parts_sheet = workbook.active
    total_parts_sheet.title = "Total Parts"
    write_to_sheet(
        total_parts_sheet,
        headers=["Description", "CPN", "MPN"],
        data=total_parts,
        fields=["description", "cpn", "mpn"]
    )

    # Add Matched Parts sheet
    matched_parts_sheet = workbook.create_sheet(title="Matched Parts")
    write_to_sheet(
        matched_parts_sheet,
        headers=["CPN", "MPN", "Manufacturer"],
        data=matched_parts,
        fields=["cpn", "mpn", "mfr"]
    )

    # Add Unmatched Parts sheet
    unmatched_parts_sheet = workbook.create_sheet(title="Unmatched Parts")
    write_to_sheet(
        unmatched_parts_sheet,
        headers=["CPN", "MPN", "Manufacturer"],
        data=unmatched_parts,
        fields=["cpn", "mpn", "mfr"]
    )

    # Save workbook to the response
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{request_id_request_type}.xlsx"'
    return response

def search_customer(request):
    query = request.GET.get('query', '')
    # Fetch all matching customers
    customers = Request.objects.filter(customer_name__icontains=query)[:5]

    # Use a dictionary to eliminate duplicates while keeping the email associated
    unique_customers = {}
    for customer in customers:
        if customer.customer_name not in unique_customers:
            unique_customers[customer.customer_name] = customer.customer_email

    # Prepare JSON response
    customer_list = [{'name': name, 'email': email} for name, email in unique_customers.items()]
    return JsonResponse(customer_list, safe=False)


def create_request(request):
    if request.method == 'POST':
        # Retrieve form data
        customer_name = request.POST['customer_name']
        customer_email = request.POST['customer_email']
        customer_code = request.POST['customer_code']
        customer_number = request.POST['customer_number']
        sales_contact = request.POST['sales_contact']
        customer_comment = request.POST['customer_comment']
        reference_code = request.POST['reference_code']
        request_type_list = request.POST.getlist('request_type')  # Allow multiple request types
        request_id = request.POST['request_id']

        # Create a single folder for the request_id to store the source email files
        request_folder = os.path.join(settings.MEDIA_ROOT, request_id)
        os.makedirs(request_folder, exist_ok=True)

        # Save the path of the first file uploaded in source_email for each Request entry
        uploaded_files = request.FILES.getlist('source_email')
        file_paths = []

        for file in uploaded_files:
            file_path = os.path.join(request_folder, file.name)
            with open(file_path, 'wb') as f:
                for chunk in file.chunks():
                    f.write(chunk)
            file_paths.append(os.path.join(request_id, file.name))  # Store relative path for saving in DB

        # Create a Request entry for each selected request type
        request_objs = []
        for req_type in request_type_list:
            req = Request(
                customer_name=customer_name,
                customer_email=customer_email,
                customer_code=customer_code,
                customer_number=customer_number,
                sales_contact=sales_contact,
                customer_comment=customer_comment,
                reference_code=reference_code,
                request_type=req_type,
                request_id=request_id,
                source_email=file_paths[0] if file_paths else None  # Store only the first file path
            )
            req.save()
            request_objs.append(req)

        # Process parts files and save data to related models for each request type
        parse_and_save_parts(request, request_id, request_objs)

        return redirect('request_list')  # Redirect to a success page or message

    return render(request, 'assapp/create_request.html')


def parse_and_save_parts(request, request_id, request_objs):
    # Helper function to parse Excel files into dictionary records
    def parse_excel(file):
        return pd.read_excel(file).to_dict(orient='records')

    # Loop through each request object and save related part data
    for req in request_objs:
        reqid_requesttype = f"{req.request_id}_{req.request_type}"
        
        # Total Parts - parse and save if file is uploaded
        if 'total_parts_file' in request.FILES:
            total_parts = parse_excel(request.FILES['total_parts_file'])
            for part in total_parts:
                TotalParts.objects.create(
                    requestid_requesttype=reqid_requesttype,
                    description=part.get('description', ''),
                    cpn=part.get('cpn', ''),
                    mpn=part.get('mpn', None)
                )
            req.total_parts_count = len(total_parts)

        # Matched Parts - parse and save if file is uploaded
        if 'matched_parts_file' in request.FILES:
            matched_parts = parse_excel(request.FILES['matched_parts_file'])
            for part in matched_parts:
                MatchedParts.objects.create(
                    requestid_requesttype=reqid_requesttype,
                    cpn=part.get('cpn', ''),
                    mpn=part.get('mpn', ''),
                    mfr=part.get('mfr', '')
                )
            req.matched_parts_count = len(matched_parts)

        # Unmatched Parts - parse and save if file is uploaded
        if 'unmatched_parts_file' in request.FILES:
            unmatched_parts = parse_excel(request.FILES['unmatched_parts_file'])
            for part in unmatched_parts:
                UnmatchedParts.objects.create(
                    requestid_requesttype=reqid_requesttype,
                    cpn=part.get('cpn', ''),
                    mpn=part.get('mpn', None)
                )
            req.unmatched_parts_count = len(unmatched_parts)

        # Save updated counts to each Request instance
        req.save()

